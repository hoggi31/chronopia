# encoding: utf-8
'''
Created on 31 mars 2016


@author: ks002450
'''

from openpyxl.reader.excel import load_workbook as _load_workbook
import contextlib

@contextlib.contextmanager
def load_workbook(*args, **kwargs):
    '''
    Open an openpyxl worksheet and automatically close it when finished.
    @see: http://stackoverflow.com/questions/31416842/openpyxl-does-not-close-excel-workbook-in-read-only-mode
    '''
    try:
        print "loaded"
        wb = _load_workbook(*args, **kwargs)
        print dir(wb)
        yield wb
    finally:
        # la faconde de fermer un workbook depend de la version de openpyxl
        pass

def ws_to_list(ws, strip=False, before_header=0, empty_col=None):
    '''
    Turns an Excel WorkSheet into a list of dict.
    '''
    rows = ws.rows #generator instanciation
    for _ in range(before_header): rows.next()
    headers = [header.value for header in rows.next()]
    if strip: headers = [(header.strip() if isinstance(header, basestring) else header) for header in headers]
    if empty_col : empty_col = headers.index(empty_col)
    return [dict(zip(headers, [cell.value for cell in row])) for row in rows if empty_col is None or row[empty_col].value is not None]

def load_wbsheets(wb_path, *sheets, **options):
    '''
    Reads a Workbook from wb_path returns a list of list of dict. 
    '''
    strip = options.get('strip', False)
    with load_workbook(wb_path, True) as wb:
        if len(sheets)==0:
            return [ws_to_list(wb.active, strip)]
        return [ws_to_list(wb.get_sheet_by_name(sheet), strip) for sheet in sheets]
