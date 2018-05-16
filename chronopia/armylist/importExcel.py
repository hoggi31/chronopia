# -*- coding: utf-8 -*-
"""
Created on Mon May 29 13:40:54 2017

@author: k005129
"""

from openpyxl import load_workbook
from armylist.models import Unit,Army,Competence,Attack, spell
print "GO"

def impSort():
   wb = load_workbook(filename='D:/Sort5.xlsx', read_only=True)
   ws = wb['Sheet2'] 
    
   headers = [header.value for header in ws.rows.next()]
   for _ in range(1):
        ws.rows.next()
    
   listCells =  [dict(zip(headers, [cell.value for cell in row])) for row in ws.rows]  
              
   for element in listCells[1:]:
        print element
        if element['armee'] is not None: 
            spellUnit,created = spell.objects.get_or_create(name= element['toto'])
            spellUnit.cost = element['cout']
            spellUnit.action = element['Actions']
            spellUnit.portee = element['Portee']
            spellUnit.level = element['Niveau']
            spellUnit.description = element['Effet']
            armyTemp = Army.objects.get(pk=element['armee']);  
            print armyTemp     
            spellUnit.army = armyTemp
            spellUnit.save()
            

def impUnit():
    wb = load_workbook(filename='D:/analyseCaracs12.xlsx', read_only=True)
    ws = wb['Sheet1']
    
    
    headers = [header.value for header in ws.rows.next()]
    for _ in range(1):
        ws.rows.next()
    
    listCells =  [dict(zip(headers, [cell.value for cell in row])) for row in ws.rows]  
              
    for element in listCells[1:]:
        print element
        
        
        if element['TOUS'] is not None: 
            unite,created = Unit.objects.get_or_create(name= element['UNITES_nom'],tir=element['Tir']);
           
            
            armyTemp = Army.objects.get(pk=element['TOUS']);       
            unite.army = armyTemp
            unite.action = element['Action']
            unite.mouvement = element['Mvt']
            unite.vol = element['Vol']
            print element['TOUS']   
            
            unite.generique = element['Generique']
            
            unite.wound = element['Y']
            unite.commandment = element['Cd']
            unite.notes = ""
            unite.defence = element['Def']
            unite.armure = element['Arm']
            unite.force = element['F']
            unite.taille = element['T']
            unite.type = element['UNITE']
            unite.type2 = element['TypeUnit']
            unite.tir = element['Tir']
            ListComp = element['Competences'].rstrip().split(',')
            
            unite.min = element['Min']
            unite.max = element['Max']
            unite.cost = element['cost']
            
            unite.save()   
            for comps in ListComp:
                persistedComp, created = Competence.objects.get_or_create(name=comps)
                unite.competences.add(persistedComp)
                print persistedComp
            
            if element['Type'] is not None: 
                 
                att1,created = Attack.objects.get_or_create(type=element['Type'], att=element['Att'],dmg=element['Dmg'],Effets=element['Effets'] )
                
                unite.attaques.add(att1.pk) 
                      
            if element['Type2'] is not None:    
                att1,created = Attack.objects.get_or_create(type=element['Type2'], att=element['Att2'],dmg=element['Dmg2'],Effets=element['Effets2'] )
                unite.attaques.add(att1.pk)
                      
            if element['Type3'] is not None: 
                att1,created =  Attack.objects.get_or_create(type=element['Type3'], att=element['Att3'],dmg=element['Dmg3'],Effets=element['Effets3'] ) 
                unite.attaques.add(att1.pk)
                                  
            if element['Type4'] is not None:  
                att1,created = Attack.objects.get_or_create(type=element['Type4'], att=element['Att4'],dmg=element['Dmg4'],Effets=element['Effets4'] )
                unite.attaques.add(att1.pk)
                                 
            if element['Type5'] is not None:
                att1,created =  Attack.objects.get_or_create(type=element['Type5'], att=element['Att5'],dmg=element['Dmg5'],Effets=element['Effets5'] ) 
                unite.attaques.add(att1.pk)
                
            
        
        
         
    
