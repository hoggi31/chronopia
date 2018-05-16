# encoding: utf-8
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook

def collectRows(wb):
    page = wb.active
    
    rows = page.iter_rows()
    header =  [(cell.values for cell in rows.Next())]#Header ? ou une chaine pr�d�finie pour creer un constructeur commun aux produits et aux composents?
    
    rowDict = [] # liste de dictionnaire : clef = Head, valuer = cellule
    for row in rows:
        rowDict.append(dict(zip(header,(cell.values for cell in page.row))))    
    
    return rowDict

def openAndLoadXlsData(path):
    #Chargement des donn�e des fichiers UIM et FING
    #UIM : extract SAP
    #Fing : extract ??
    UIM = Workbook()
    xlsData = load_workbook(path)
    #Creation d'une liste de dictionnaire  
    #Dictionnaire: clef = entetes : valeur = valeur des cellules
    xlsData_index = collectRows(xlsData);
    
    # construction du set de produit
    objects = set()
    for row in  xlsData_index:
       objects.add(Components(row))   

    #TODO: En fonction de l'emplacement de l'information d'association charger l'un ou l'autre en premier   
    pass


class Components(Objects):
    '''
    classdocs
    '''
    def __init__(self, params):
        self.component_name = params['name']
        self.id = params['id']
        self.part_number = params['pn']
        self.isReach = True
        self.parentProducts = []
    
    def addParentProduct(self, parent):
        self.parentProduct.append(parent)   
        